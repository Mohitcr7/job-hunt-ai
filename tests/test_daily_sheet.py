# tests/test_daily_sheet.py
#
# Unit tests for the daily spreadsheet: the field normalisers that turn raw
# scraper output into readable cells, the fit banding, and the sheet writer.
# No network and no LLM — the writer is fed hand-built JobMatch objects.

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def _job(**overrides):
    from tools.vector_store_tool import Job
    fields = dict(
        title="AI/ML Engineer", company="Acme", location="Bengaluru",
        description="x", url="https://example.test/1", platform="linkedin",
        posted_date="2026-07-29 04:51:00", salary="", experience="", job_type="",
    )
    fields.update(overrides)
    return Job(**fields)


def _match(score, **overrides):
    from tools.vector_store_tool import JobMatch
    return JobMatch(job=_job(**overrides), score=score / 100, score_percent=score)


# --- field normalisers ------------------------------------------------------

def test_experience_is_title_cased():
    from tools.scraper_tool import _pretty_experience
    assert _pretty_experience("mid-senior level") == "Mid-senior level"
    assert _pretty_experience("entry level") == "Entry level"


def test_not_applicable_experience_becomes_empty():
    """LinkedIn says 'not applicable' where it means 'unknown'."""
    from tools.scraper_tool import _pretty_experience
    assert _pretty_experience("not applicable") == ""
    assert _pretty_experience("nan") == ""
    assert _pretty_experience(None) == ""


def test_job_type_slug_becomes_readable():
    from tools.scraper_tool import _pretty_job_type
    assert _pretty_job_type("fulltime") == "Full-time"
    assert _pretty_job_type("parttime") == "Part-time"
    # Multi-valued: show the first, don't render the raw comma-joined slug.
    assert _pretty_job_type("contract,temporary") == "Contract"


def test_salary_range_formats_both_bounds():
    from tools.scraper_tool import _salary_range
    assert _salary_range(1500000, 2500000, "INR", "yearly") == "₹1,500,000–₹2,500,000/yearly"


def test_undisclosed_salary_is_blank_not_nan():
    """pandas hands us NaN for missing pay; 'nan' must never reach a cell."""
    from tools.scraper_tool import _salary_range
    assert _salary_range("nan", "nan", "nan", "nan") == ""
    assert _salary_range(None, None, None, None) == ""


# --- fit banding ------------------------------------------------------------

def test_bands_split_at_the_configured_cutoffs():
    from exports.daily_sheet import _band, STRONG_BAND, REACHABLE_BAND
    assert _band(STRONG_BAND, reranked=True) == "Strong"
    assert _band(STRONG_BAND - 1, reranked=True) == "Reachable"
    assert _band(REACHABLE_BAND, reranked=True) == "Reachable"
    assert _band(REACHABLE_BAND - 1, reranked=True) == "Stretch"


def test_each_scorer_uses_its_own_band_scale():
    """
    Reranked runs score by percentile, embedding runs by cosine similarity.
    The scales differ, so the same number must not mean the same band — 70 is
    mid-pack among percentiles but a high cosine score.
    """
    from exports.daily_sheet import _band
    assert _band(70, reranked=True) == "Reachable"
    assert _band(70, reranked=False) == "Strong"


def test_posted_time_converts_utc_to_ist():
    """04:51 UTC is 10:21 IST — a sheet labelled IST must actually be IST."""
    from exports.daily_sheet import _posted_ist
    label, _ = _posted_ist(_job(posted_date="2026-07-29 04:51:00"))
    assert label == "29 Jul 10:21"


def test_undated_rows_sort_last_rather_than_first():
    from exports.daily_sheet import _posted_ist
    label, key = _posted_ist(_job(posted_date=""))
    assert label == "—"
    assert key.year == 1970


# --- deduplication ----------------------------------------------------------

def test_same_role_from_several_searches_is_collapsed(monkeypatch):
    """One opening surfaced by N searches gets N URLs — it's still one job."""
    import tools.scraper_tool as st

    same_role = [
        _job(url=f"https://example.test/{i}", title="AI / ML Engineer", company="Accenture")
        for i in range(5)
    ]
    other = _job(url="https://example.test/x", title="Data Scientist", company="Acme")

    monkeypatch.setattr(st, "scrape_with_jobspy", lambda *a, **k: same_role + [other])
    result = st.scrape_all_jobs(search_terms=["x"], locations=["y"], enable_naukri=False)

    assert len(result) == 2
    assert {j.title for j in result} == {"AI / ML Engineer", "Data Scientist"}


def test_untitled_jobs_are_not_collapsed_together(monkeypatch):
    """Blank title+company must not make every such row look like a duplicate."""
    import tools.scraper_tool as st

    jobs = [_job(url=f"https://example.test/{i}", title="", company="") for i in range(3)]
    monkeypatch.setattr(st, "scrape_with_jobspy", lambda *a, **k: jobs)
    result = st.scrape_all_jobs(search_terms=["x"], locations=["y"], enable_naukri=False)

    assert len(result) == 3


# --- sheet writer -----------------------------------------------------------

def test_sheet_layout_and_hyperlinks(tmp_path):
    from openpyxl import load_workbook
    from exports.daily_sheet import write_sheet

    matches = [
        _match(95, title="AI/ML Engineer", experience="1-4 yrs",
               salary="₹20-30 LPA", job_type="Full-time", platform="naukri"),
        _match(65, title="Remote Data Scientist", job_type="Contract"),
        _match(20, title="Intern", url="", platform="company_page"),
    ]
    path = write_sheet(matches, str(tmp_path))
    sheet = load_workbook(path).active

    assert sheet.freeze_panes == "A5"           # headers stay visible while scrolling
    assert sheet.auto_filter.ref == "A4:K7"
    assert sheet.cell(row=4, column=1).value == "#"

    assert sheet.cell(row=5, column=2).value == "Naukri"
    assert sheet.cell(row=5, column=10).value == "Strong"
    assert sheet.cell(row=5, column=11).hyperlink.target == "https://example.test/1"

    # A job with no link gets a dash, not a broken hyperlink.
    assert sheet.cell(row=7, column=11).value == "—"
    assert sheet.cell(row=7, column=11).hyperlink is None


def test_missing_fields_render_as_dashes(tmp_path):
    from openpyxl import load_workbook
    from exports.daily_sheet import write_sheet

    path = write_sheet([_match(50, experience="", salary="", job_type="")], str(tmp_path))
    sheet = load_workbook(path).active
    for column in (7, 8, 9):
        assert sheet.cell(row=5, column=column).value == "—"


def test_filename_is_dated(tmp_path):
    from datetime import datetime
    from exports.daily_sheet import write_sheet, IST

    path = write_sheet([_match(50)], str(tmp_path))
    assert path.name == f"Jobs-{datetime.now(IST):%Y-%m-%d}.xlsx"


def test_empty_scrape_still_writes_a_sheet(tmp_path):
    """The writer itself stays total — an empty list is a readable sheet."""
    from openpyxl import load_workbook
    from exports.daily_sheet import write_sheet

    path = write_sheet([], str(tmp_path))
    sheet = load_workbook(path).active
    assert "0 roles" in sheet["A2"].value
    assert sheet.cell(row=4, column=1).value == "#"


# --- an empty scrape is a failed run, not a quiet day -----------------------

def test_empty_scrape_raises_instead_of_writing(tmp_path, monkeypatch):
    import exports.daily_sheet as ds

    monkeypatch.setattr(ds, "ALLOW_EMPTY", False)
    monkeypatch.setattr(ds, "collect_jobs", lambda **kwargs: [])

    with pytest.raises(ds.EmptyScrapeError):
        ds.build_daily_sheet(output_dir=str(tmp_path))


def test_empty_scrape_does_not_clobber_an_existing_sheet(tmp_path, monkeypatch):
    """
    The regression that made this worth guarding: the Mac lost DNS at noon on
    2026-08-08, every source failed, and a good sheet was replaced by a header
    row. Yesterday's data is worth more than an empty file.
    """
    import exports.daily_sheet as ds

    good = ds.write_sheet([_match(80)], str(tmp_path))
    before = good.read_bytes()

    monkeypatch.setattr(ds, "ALLOW_EMPTY", False)
    monkeypatch.setattr(ds, "collect_jobs", lambda **kwargs: [])
    with pytest.raises(ds.EmptyScrapeError):
        ds.build_daily_sheet(output_dir=str(tmp_path))

    assert good.read_bytes() == before


def test_allow_empty_opts_back_into_writing(tmp_path, monkeypatch):
    """A genuinely narrow search can still ask for the empty sheet."""
    import exports.daily_sheet as ds

    monkeypatch.setattr(ds, "ALLOW_EMPTY", True)
    monkeypatch.setattr(ds, "collect_jobs", lambda **kwargs: [])

    path = ds.build_daily_sheet(output_dir=str(tmp_path))
    assert path.exists()


# --- cross-encoder reranking ------------------------------------------------

def _fake_rerank(monkeypatch, scores_by_batch):
    """Stub the rerank endpoint, returning caller-supplied relevance scores."""
    import tools.reranker_tool as rr
    from types import SimpleNamespace

    calls = []

    def post(url, **kwargs):
        documents = kwargs["json"]["documents"]
        calls.append(len(documents))
        scores = scores_by_batch(len(calls) - 1, len(documents))
        return SimpleNamespace(
            status_code=200,
            raise_for_status=lambda: None,
            json=lambda: {"results": [
                {"index": i, "relevance_score": s} for i, s in enumerate(scores)
            ]},
        )

    monkeypatch.setattr(rr, "OPENROUTER_API_KEY", "test-key")
    monkeypatch.setitem(sys.modules, "requests", SimpleNamespace(post=post))
    return calls


def test_reranker_reorders_by_relevance(monkeypatch):
    from tools.reranker_tool import score_all_jobs

    # Embedding order is 1,2,3; the cross-encoder disagrees and prefers the last.
    matches = [_match(60, title="A"), _match(50, title="B"), _match(40, title="C")]
    _fake_rerank(monkeypatch, lambda batch, n: [0.001, 0.002, 0.009])

    result = score_all_jobs("resume text", matches)
    assert [m.job.title for m in result] == ["C", "B", "A"]


def test_tiny_relevance_scores_still_produce_a_usable_spread(monkeypatch):
    """
    The model returns relevance in the 0.0003-0.015 range on real postings.
    Multiplying by 100 would collapse every row to 0 or 1, so scores must be
    restated as percentiles instead.
    """
    from tools.reranker_tool import score_all_jobs

    matches = [_match(50, title=f"J{i}") for i in range(10)]
    _fake_rerank(monkeypatch, lambda batch, n: [0.0001 * (n - i) for i in range(n)])

    result = score_all_jobs("resume text", matches)
    percents = [m.score_percent for m in result]
    assert percents[0] == 100
    assert len(set(percents)) > 2, "scores collapsed instead of spreading"
    assert percents == sorted(percents, reverse=True)


def test_large_batches_are_chunked_under_the_api_limit(monkeypatch):
    """The endpoint rejects >512 passages: 'List should have at most 512 items'."""
    from tools.reranker_tool import score_all_jobs, RERANK_BATCH_SIZE

    matches = [_match(50, title=f"J{i}") for i in range(900)]
    calls = _fake_rerank(monkeypatch, lambda batch, n: [0.001] * n)

    score_all_jobs("resume text", matches)
    assert len(calls) > 1, "900 jobs must not go in one request"
    assert max(calls) <= RERANK_BATCH_SIZE <= 512


def test_rerank_failure_keeps_embedding_order(monkeypatch):
    """A rate limit must degrade ranking quality, not lose the run."""
    from types import SimpleNamespace
    import tools.reranker_tool as rr
    from tools.reranker_tool import score_all_jobs

    def boom(*a, **k):
        raise OSError("429 rate limited")

    monkeypatch.setattr(rr, "OPENROUTER_API_KEY", "test-key")
    monkeypatch.setitem(sys.modules, "requests", SimpleNamespace(post=boom))

    matches = [_match(60, title="A"), _match(50, title="B")]
    result = score_all_jobs("resume text", matches)
    assert [m.job.title for m in result] == ["A", "B"]
    assert result[0].score_percent == 60


def test_no_api_key_keeps_embedding_order(monkeypatch):
    import tools.reranker_tool as rr
    from tools.reranker_tool import score_all_jobs

    monkeypatch.setattr(rr, "OPENROUTER_API_KEY", "")
    matches = [_match(60, title="A"), _match(50, title="B")]
    assert [m.job.title for m in score_all_jobs("q", matches)] == ["A", "B"]
