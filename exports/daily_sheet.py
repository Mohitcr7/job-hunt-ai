# exports/daily_sheet.py
#
# WHAT THIS FILE DOES:
# Produces the daily job spreadsheet — one .xlsx per day listing every freshly
# scraped posting with a fit band, ready to skim over coffee.
#
# WHY IT ISN'T THE FULL PIPELINE:
# `python main.py run` tailors a resume and cover letter for each match, which
# costs an LLM call per job. A daily sheet covers *everything* scraped (often
# 100+ roles), so tailoring them all would be slow and expensive for no benefit
# — you only want kits for the handful you actually apply to.
#
# So this ranks with the free half of the matcher: local FAISS embeddings score
# every job, and the LLM only gets involved if you ask for it (--llm-top N).
# That keeps the daily run free and offline apart from the scrape itself.

import os
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Optional

from loguru import logger

from tools.vector_store_tool import Job, JobMatch

# Sharpen the top of the sheet with the cross-encoder reranker. Free (separate
# budget from the generation LLM) and ~11s for a day's scrape.
ENABLE_SHEET_RERANK = os.getenv("ENABLE_SHEET_RERANK", "true").lower() in ("1", "true", "yes")

# How many of FAISS's best go to the reranker. Everything below stays in
# embedding order — see the note in collect_jobs for why the split exists.
RERANK_POOL = int(os.getenv("SHEET_RERANK_POOL", "200"))

# Fit bands. The two scorers put numbers on different scales, so each gets its
# own cutoffs. Reranked runs score a job by its PERCENTILE within the day's
# batch, so 90 means "top 10% of what was scraped today" — relative to the day,
# not an absolute verdict. Embedding runs score raw cosine similarity, which is
# absolute but much blunter. Neither is the LLM's judgement.
STRONG_BAND = int(os.getenv("SHEET_STRONG_BAND", "90"))
REACHABLE_BAND = int(os.getenv("SHEET_REACHABLE_BAND", "60"))
FAISS_STRONG_BAND = int(os.getenv("SHEET_FAISS_STRONG_BAND", "55"))
FAISS_REACHABLE_BAND = int(os.getenv("SHEET_FAISS_REACHABLE_BAND", "42"))

# A scrape that returns nothing is treated as a failed run, not a quiet day.
# Set this if you genuinely expect empty results (a very narrow search, say).
ALLOW_EMPTY = os.getenv("SHEET_ALLOW_EMPTY", "false").lower() in ("1", "true", "yes")

IST = timezone(timedelta(hours=5, minutes=30))


class EmptyScrapeError(RuntimeError):
    """
    Every source returned zero jobs.

    Raised instead of writing the sheet, because an empty result is far more
    often a broken run than an empty job market — on 2026-08-08 the Mac lost
    DNS at noon, every source failed, and the run cheerfully overwrote a good
    spreadsheet with a header row and exited 0. Failing loudly keeps yesterday's
    sheet intact and lets the wrapper script raise a desktop notification.
    """

COLUMNS = [
    ("#", 5),
    ("Source", 11),
    ("Job Title", 42),
    ("Company", 26),
    ("Location", 26),
    ("Posted (IST)", 17),
    ("Experience", 13),
    ("Salary", 16),
    ("Type", 12),
    ("Fit", 12),
    ("Apply Link", 15),
]

# Sheet palette, picked to stay legible in Excel, Numbers and Google Sheets.
NAVY = "1F3864"
INK = "1F3864"
MUTED = "7F7F7F"
BANDS = {
    #            fill      font
    "Strong":    ("D9EAD3", "1E6B34"),
    "Reachable": ("FFF2CC", "8A6100"),
    "Stretch":   ("F2F2F2", "5A5A5A"),
}

SOURCE_LABELS = {
    "linkedin": "LinkedIn",
    "indeed": "Indeed",
    "naukri": "Naukri",
    "company_page": "Company",
}


def _band(score_percent: int, reranked: bool = None) -> str:
    if reranked is None:
        reranked = ENABLE_SHEET_RERANK
    strong = STRONG_BAND if reranked else FAISS_STRONG_BAND
    reachable = REACHABLE_BAND if reranked else FAISS_REACHABLE_BAND
    if score_percent >= strong:
        return "Strong"
    if score_percent >= reachable:
        return "Reachable"
    return "Stretch"


def _posted_ist(job: Job) -> tuple:
    """
    Render the posting time in IST, and return a sort key alongside it.

    Sources disagree about precision: jobspy may give a full timestamp, a bare
    date, or nothing at all. The sort key falls back to the epoch so undated
    rows sink rather than jumping to the top.
    """
    raw = (job.posted_date or "").strip()
    if not raw:
        return "—", datetime.fromtimestamp(0, tz=IST)

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(raw[:len(fmt) + 2].strip(), fmt)
        except ValueError:
            continue
        # Naive timestamps from the boards are UTC; dates carry no time at all.
        parsed = parsed.replace(tzinfo=timezone.utc).astimezone(IST)
        label = parsed.strftime("%d %b") if fmt == "%Y-%m-%d" else parsed.strftime("%d %b %H:%M")
        return label, parsed

    return raw, datetime.fromtimestamp(0, tz=IST)


def collect_jobs(
    search_terms: Optional[List[str]] = None,
    locations: Optional[List[str]] = None,
    hours_old: int = 24,
    resume_path: Optional[str] = None,
    llm_top_n: int = 0,
) -> List[JobMatch]:
    """
    Scrape, score every job against the resume, and return them sorted by fit
    then recency. `llm_top_n` optionally re-scores that many top rows with the
    real LLM fit-scorer — off by default so the daily run stays free.
    """
    from agents.scout_agent import get_resume
    from agents.matcher_agent import faiss_prefilter
    from tools.scraper_tool import scrape_all_jobs
    from config import JOB_PREFERENCES

    resume = get_resume(resume_path or "data/resume.pdf")
    jobs = scrape_all_jobs(
        search_terms=search_terms,
        locations=locations,
        hours_old=hours_old,
    )
    if not jobs:
        logger.warning("Scrape returned no jobs from any source.")
        return []

    # min_score=0 and top_k=len(jobs): unlike the pipeline we want a score for
    # every posting, not a shortlist. Embedding all of them is local and free.
    matches = faiss_prefilter(resume, jobs, min_score=0, top_k=len(jobs))

    # Rerank the top slice, not everything. Measured against planted controls in
    # a real 650-job scrape: the reranker ordered the top better than FAISS
    # (a perfect-match role rose 16 → 8) but was far worse at burying obvious
    # noise — it put a chartered-accountancy job at 237/654 where FAISS had it
    # at 652. So FAISS does the rejecting it's reliably good at, and the
    # reranker only re-orders a pool already free of nurses and accountants.
    matches.sort(key=lambda m: (m.score_percent, _posted_ist(m.job)[1]), reverse=True)

    if ENABLE_SHEET_RERANK and len(matches) > 1:
        from tools.reranker_tool import score_all_jobs
        query = f"{resume.summary} Skills: {', '.join(resume.skills)}"
        pool, tail = matches[:RERANK_POOL], matches[RERANK_POOL:]
        matches = score_all_jobs(query, pool) + tail

        # The pool now carries cross-encoder scores and the tail still carries
        # cosine ones — two different scales that must not be compared. Ordering
        # is already correct, so restate every row as its percentile within the
        # day's batch. That is also what the Fit bands are calibrated against.
        total = len(matches)
        for position, match in enumerate(matches):
            match.score_percent = int(round(100 * (total - position) / total))

    if llm_top_n > 0:
        matches = _llm_rescore_top(resume, matches, llm_top_n,
                                   JOB_PREFERENCES.get("experience_years", 0))
        matches.sort(key=lambda m: (m.score_percent, _posted_ist(m.job)[1]), reverse=True)

    return matches


def _llm_rescore_top(resume, matches: List[JobMatch], top_n: int, experience_years: int):
    """
    Replace the embedding score with a real LLM fit score for the top N rows.

    Opt-in: it costs one call per row. Any row whose call fails silently keeps
    its embedding score, so a rate limit degrades the ranking instead of the run.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from agents.matcher_agent import llm_score_job
    from config import LLM_CONCURRENCY

    top = matches[:top_n]
    logger.info(f"LLM-scoring the top {len(top)} of {len(matches)} jobs")

    with ThreadPoolExecutor(max_workers=LLM_CONCURRENCY) as pool:
        futures = {
            pool.submit(llm_score_job, resume, m.job, experience_years): m
            for m in top
        }
        for future in as_completed(futures):
            match = futures[future]
            try:
                match.score_percent = future.result().score
                match.score = match.score_percent / 100
            except Exception as e:
                logger.debug(f"LLM scoring failed for {match.job.title}: {e}")

    return matches


def write_sheet(
    matches: List[JobMatch],
    output_dir: str,
    hours_old: int = 24,
    scored_by: str = "embedding similarity",
) -> Path:
    """Write the .xlsx and return its path. Filename is the run's date."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    now = datetime.now(IST)
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    path = destination / f"Jobs-{now:%Y-%m-%d}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = now.strftime("%d %b %Y")

    roles = os.getenv("SHEET_TITLE", "Job matches")
    sheet["A1"] = f"{roles} — posted in the last {hours_old} hours"
    sheet["A1"].font = Font(bold=True, size=14, color=INK)

    sheet["A2"] = (
        f"Scraped {now:%d %b %Y, %H:%M IST} • postings from the last {hours_old} hours only "
        f"• {len(matches)} roles • sorted by fit, then recency • fit from {scored_by}"
    )
    sheet["A2"].font = Font(italic=True, size=9, color=MUTED)

    header_row = 4
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")

    for index, (label, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[header_row].height = 20

    for offset, match in enumerate(matches):
        row = header_row + 1 + offset
        job = match.job
        posted_label, _ = _posted_ist(job)
        band = _band(match.score_percent)

        values = [
            offset + 1,
            SOURCE_LABELS.get(job.platform, (job.platform or "—").title()),
            job.title or "—",
            job.company or "—",
            job.location or "—",
            posted_label,
            job.experience or "—",
            job.salary or "—",
            job.job_type or "—",
            band,
            None,   # hyperlink, set below
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if index == 1 else "left")
            cell.font = Font(size=10)

        fit_cell = sheet.cell(row=row, column=10)
        fill_colour, font_colour = BANDS[band]
        fit_cell.fill = PatternFill("solid", fgColor=fill_colour)
        fit_cell.font = Font(bold=True, size=10, color=font_colour)

        link_cell = sheet.cell(row=row, column=11)
        if job.url:
            link_cell.value = "Open posting"
            link_cell.hyperlink = job.url
            link_cell.font = Font(color="0563C1", underline="single", size=10)
        else:
            link_cell.value = "—"

    last_row = header_row + len(matches)
    if matches:
        sheet.auto_filter.ref = f"A{header_row}:K{last_row}"
    sheet.freeze_panes = f"A{header_row + 1}"

    workbook.save(path)
    logger.info(f"Wrote {len(matches)} rows to {path}")
    return path


def build_daily_sheet(
    output_dir: str,
    search_terms: Optional[List[str]] = None,
    locations: Optional[List[str]] = None,
    hours_old: int = 24,
    resume_path: Optional[str] = None,
    llm_top_n: int = 0,
) -> Path:
    """
    Scrape, score and write today's sheet. Returns the file path.

    Raises EmptyScrapeError if nothing was scraped, leaving any existing sheet
    for the day untouched — see that class for why an empty run is a failure.
    """
    matches = collect_jobs(
        search_terms=search_terms,
        locations=locations,
        hours_old=hours_old,
        resume_path=resume_path,
        llm_top_n=llm_top_n,
    )
    if not matches and not ALLOW_EMPTY:
        raise EmptyScrapeError(
            "Every source returned zero jobs. That is almost always a network "
            "outage or a broken scraper rather than an empty job market, so the "
            "sheet was left as it was instead of being overwritten with an empty "
            "one. Check logs/daily-sheet.log for the per-source errors, or set "
            "SHEET_ALLOW_EMPTY=true if an empty result is genuinely expected."
        )

    scored_by = "cross-encoder reranker" if ENABLE_SHEET_RERANK else "embedding similarity"
    if llm_top_n > 0:
        scored_by = f"LLM fit scoring (top {llm_top_n}) + {scored_by}"
    return write_sheet(matches, output_dir, hours_old=hours_old, scored_by=scored_by)
